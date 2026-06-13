#!/usr/bin/env python3
"""Import an IATA 'Monthly Air Traffic Data Detail (YoY)' workbook into a
compact JSON the dashboard can read. (Source: IATA, iata.org.)

The client's workbook has three views (sheets): System, International, Domestic.
Each view is a wide month-by-month grid: column GROUPS = Industry + regions
(Domestic = countries), each with metrics RPK, ASK, FTK, ATK, PLF, plus
'3mn rolling' copies and RPK/FTK market-share blocks.

We keep only the single-month blocks (the rolling averages are derived in the
browser) plus RPK market share. Values are stored as percentages, 1 decimal:
  - RPK/ASK/FTK/ATK  -> YoY % change   (0.119  -> 11.9)
  - PLF              -> load factor %  (0.774  -> 77.4)
  - share            -> % of view RPK  (0.182  -> 18.2)
"""
from __future__ import annotations
import datetime as dt, json, os, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "iata_detail.json")
# RPK/ASK/PLF are populated across the whole window; the workbook's freight
# columns (FTK/ATK) are blank for recent years, so we mirror the client's
# populated trio plus RPK market share.
METRICS = ["RPK", "ASK", "PLF"]
START = "2011-04"           # align all three views to a common 15-year window
SHEETS = ["System", "International", "Domestic"]


def ym(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return "%04d-%02d" % (v.year, v.month)
    return None


def pct(v):
    if v is None or isinstance(v, str):
        return None
    try:
        return round(float(v) * 100, 1)
    except (TypeError, ValueError):
        return None


def forward_fill(row):
    out, cur = [], ""
    for c in row:
        s = (str(c).replace("\n", " ").strip() if c is not None else "")
        if s:
            cur = s
        out.append(cur)
    return out


def is_share_group(name, subs):
    if "market share" in name.lower():
        return True
    # International uses bare 'RPK' / 'FTK' group names for its share blocks
    return name.strip() in ("RPK", "FTK") and ("Total" in subs or "Asia Pacific" in subs)


def is_rpk_share(name):
    n = name.strip().lower()
    return n == "rpk" or "rpk (market share" in n or "rpk(market share" in n


def clean_group(name):
    return name.split("(")[0].strip()        # drop '(3mn rolling)' / '(rolling)'


def import_sheet(ws):
    row1 = [c.value for c in ws[1]]
    row2 = [c.value for c in ws[2]]
    groups = forward_fill(row1)
    subs = ["" if c is None else str(c).strip() for c in row2]
    ncol = max(len(groups), len(subs))
    groups += [""] * (ncol - len(groups))
    subs += [""] * (ncol - len(subs))

    # month index (col A, rows 4..end), trimmed to the common window
    months, rows = [], []
    for r in range(4, ws.max_row + 1):
        m = ym(ws.cell(row=r, column=1).value)
        if m and m >= START:
            months.append(m)
            rows.append(r)

    # column spans per group, in sheet order
    spans, order = {}, []
    for i, g in enumerate(groups):
        if not g:
            continue
        spans.setdefault(g, [None, None])
        if spans[g][0] is None:
            spans[g][0] = i + 1
            order.append(g)
        spans[g][1] = i + 1

    def col_values(ci):
        return [pct(ws.cell(row=r, column=ci).value) for r in rows]

    data_groups, share = {}, {}
    group_order = []
    for g in order:
        a, b = spans[g]
        grp_subs = [subs[j] for j in range(a - 1, b)]
        if "rolling" in g.lower():
            continue                                   # derived in the browser
        if is_share_group(g, grp_subs):
            if not is_rpk_share(g):
                continue                               # keep RPK share, skip FTK share
            for j in range(a - 1, b):
                nm = subs[j].strip()
                if nm:
                    vals = col_values(j + 1)
                    if any(x is not None for x in vals):
                        share[nm] = vals
            continue
        # data group: take the FIRST column of each metric within the span
        series = {}
        for met in METRICS:
            for j in range(a - 1, b):
                if subs[j].strip().upper() == met and met not in series:
                    series[met.lower()] = col_values(j + 1)
                    break
        # skip near-empty groups (e.g. an unused 'China' block in System with a
        # single stray value); a real series has many months populated
        nn = max((sum(x is not None for x in s) for s in series.values()), default=0)
        if series and nn >= 12:
            cg = clean_group(g)
            if cg not in data_groups:
                data_groups[cg] = series
                group_order.append(cg)

    return {"months": months, "order": group_order,
            "groups": data_groups, "share_rpk": share}


def main():
    src = sys.argv[1]
    wb = openpyxl.load_workbook(src, data_only=True)
    views = {}
    for s in SHEETS:
        if s in wb.sheetnames:
            views[s.lower()] = import_sheet(wb[s])

    latest = max(v["months"][-1] for v in views.values())
    out = {
        "_meta": {
            "source": "IATA Air Passenger Market Analysis — iata.org "
                      "(monthly regional detail, year-on-year).",
            "generated": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "latest": latest,
            "window_from": START,
            "metrics": {"rpk": "Passenger traffic (RPK), YoY %",
                        "ask": "Passenger capacity (ASK), YoY %",
                        "plf": "Passenger load factor, %",
                        "ftk": "Cargo traffic (FTK), YoY %",
                        "atk": "Cargo capacity (ATK), YoY %"},
        },
        "views": views,
    }
    with open(OUT, "w") as f:
        json.dump(out, f, separators=(",", ":"))

    print("wrote", OUT)
    for name, v in views.items():
        print("  %-13s %s..%s  groups: %s"
              % (name, v["months"][0], v["months"][-1], ", ".join(v["order"])))


if __name__ == "__main__":
    raise SystemExit(main())
